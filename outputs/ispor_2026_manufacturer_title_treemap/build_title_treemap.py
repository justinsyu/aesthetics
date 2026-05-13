from collections import Counter
from html import escape
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Users/justinyu/Desktop/linkedin-posts")
INPUT = Path("/Users/justinyu/Downloads/ISPOR_2026_attendees (2).xlsx")
OUT_DIR = ROOT / "outputs" / "ispor_2026_manufacturer_title_treemap"
HTML_OUT = OUT_DIR / "ispor_2026_manufacturer_title_treemap.html"
CSV_OUT = OUT_DIR / "manufacturer_title_treemap_counts.csv"

SHEET_W = 1200
SHEET_H = 2600
TREEMAP_W = 1048
TREEMAP_H = 1960
GAP = 6


def pct(value, total):
    return f"{round(value / total * 100)}%" if total else "0%"


def read_counts():
    wb = load_workbook(INPUT, read_only=True, data_only=True)
    ws = wb["Attendees"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    title_idx = headers.index("Normalized Title")
    affiliation_type_idx = headers.index("Affiliation Type")

    counts = Counter()
    blank = 0
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[affiliation_type_idx] or "").strip() != "Manufacturer":
            continue
        total += 1
        title = str(row[title_idx] or "").strip()
        if not title or title.lower() in {"nan", "none"}:
            blank += 1
            continue
        counts[title] += 1
    return total, blank, counts


def normalize_sizes(sizes, width, height):
    area = width * height
    total = sum(sizes)
    return [size * area / total for size in sizes]


def worst_ratio(row, side):
    if not row:
        return float("inf")
    row_sum = sum(row)
    row_min = min(row)
    row_max = max(row)
    side_sq = side * side
    return max((side_sq * row_max) / (row_sum * row_sum), (row_sum * row_sum) / (side_sq * row_min))


def layout_row(row, x, y, width, height):
    rects = []
    row_area = sum(row)
    if width >= height:
        row_width = row_area / height
        cy = y
        for size in row:
            rect_h = size / row_width
            rects.append({"x": x, "y": cy, "dx": row_width, "dy": rect_h})
            cy += rect_h
        return rects, x + row_width, y, width - row_width, height

    row_height = row_area / width
    cx = x
    for size in row:
        rect_w = size / row_height
        rects.append({"x": cx, "y": y, "dx": rect_w, "dy": row_height})
        cx += rect_w
    return rects, x, y + row_height, width, height - row_height


def squarify(sizes, x, y, width, height):
    sizes = list(sizes)
    rects = []
    while sizes:
        row = []
        side = min(width, height)
        while sizes:
            candidate = row + [sizes[0]]
            if not row or worst_ratio(candidate, side) <= worst_ratio(row, side):
                row.append(sizes.pop(0))
            else:
                break
        new_rects, x, y, width, height = layout_row(row, x, y, width, height)
        rects.extend(new_rects)
    return rects


def write_csv(items, nonblank):
    lines = ["rank,normalized_title,count,share_of_populated_manufacturer_titles"]
    for rank, (title, count) in enumerate(items, 1):
        lines.append(
            f'{rank},"{title.replace(chr(34), chr(34) * 2)}",{count},{count / nonblank:.6f}'
        )
    CSV_OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")


def color_for_rank(rank):
    if rank <= 3:
        return "#d7ff5f", "#10120f", "rgba(16,18,15,.42)"
    if rank <= 9:
        return "#ffb86b", "#10120f", "rgba(16,18,15,.42)"
    if rank <= 20:
        return "#b8d8ff", "#10120f", "rgba(16,18,15,.42)"
    if rank <= 35:
        return "#ffd3e0", "#10120f", "rgba(16,18,15,.42)"
    return "#72746c", "#f6f1e8", "rgba(246,241,232,.22)"


def tile_class(width, height, count):
    area = width * height
    if count <= 2 or area < 5200 or min(width, height) < 44:
        return "count"
    if area >= 32000 and min(width, height) >= 95:
        return "large"
    if area >= 14000 and min(width, height) >= 66:
        return "medium"
    return "small"


def build_tiles(items):
    sizes = normalize_sizes([count for _, count in items], TREEMAP_W - 2 * GAP, TREEMAP_H - 2 * GAP)
    rects = squarify(sizes, GAP / 2, GAP / 2, TREEMAP_W - GAP, TREEMAP_H - GAP)
    html = []
    for rank, ((title, count), rect) in enumerate(zip(items, rects), 1):
        left = rect["x"] + GAP / 2
        top = rect["y"] + GAP / 2
        width = max(1, rect["dx"] - GAP)
        height = max(1, rect["dy"] - GAP)
        bg, color, border = color_for_rank(rank)
        cls = tile_class(width, height, count)
        title_attr = escape(f"{title}: {count}")
        style = (
            f"left:{left:.3f}px;top:{top:.3f}px;width:{width:.3f}px;height:{height:.3f}px;"
            f"background:{bg};color:{color};border-color:{border};"
        )
        rank_class = " tier-top" if rank <= 3 else " tier-strong" if rank <= 9 else ""
        if cls == "count":
            body = f'<div class="count-only"><span>{count}</span></div>'
        else:
            noun = "row" if count == 1 else "rows"
            body = (
                '<div class="tile-copy">'
                f'<div class="tile-rank">{rank:02d}</div>'
                f'<div class="tile-name">{escape(title)}</div>'
                f'<div class="tile-count">{count} {noun}</div>'
                "</div>"
            )
        html.append(
            f'          <div class="tile {cls}{rank_class}" title="{title_attr}" style="{style}">{body}</div>'
        )
    return "\n".join(html)


def build_html(total, blank, counts):
    items = counts.most_common()
    nonblank = sum(counts.values())
    top20_total = sum(count for _, count in items[:20])
    director_type = sum(
        counts[title]
        for title in [
            "Director",
            "Senior Director",
            "Associate Director",
            "Executive Director",
        ]
    )
    small_titles = sum(1 for _, count in items if count <= 4)
    tiles = build_tiles(items)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>ISPOR 2026 Manufacturer Title Treemap</title>
  <style>
    :root {{ --ink:#10120f; --muted:#5c6257; --paper:#f6f1e8; --card:#fffaf0; --line:#1b1f17; --lime:#d7ff5f; --orange:#ffb86b; --blue:#b8d8ff; --pink:#ffd3e0; --gray:#d6d0c2; --radius:26px; }}
    * {{ box-sizing:border-box; }} html,body {{ margin:0; background:var(--paper); scrollbar-width:none; }} html::-webkit-scrollbar, body::-webkit-scrollbar {{ display:none; }}
    body {{ color:var(--ink); font-family:Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; line-height:1.28; }}
    .sheet {{ width:{SHEET_W}px; height:{SHEET_H}px; padding:62px 52px 52px; overflow:hidden; background:var(--paper); position:relative; }}
    .top {{ display:grid; grid-template-columns:520px 1fr; gap:32px; align-items:end; }}
    .eyebrow {{ display:inline-flex; align-items:center; width:fit-content; border:1.4px solid var(--line); padding:8px 11px; border-radius:999px; font-size:15px; font-weight:850; letter-spacing:.06em; text-transform:uppercase; background:var(--lime); margin-bottom:18px; }}
    h1 {{ margin:0; max-width:520px; font-size:66px; line-height:.9; letter-spacing:-0.03em; font-weight:540; }}
    .metrics {{ display:grid; grid-template-columns:repeat(4,1fr); gap:12px; }} .metric {{ min-height:112px; border:1.5px solid var(--line); border-radius:var(--radius); background:var(--card); padding:15px; }} .metric.dark {{ background:#11130f; color:var(--paper); }}
    .num {{ font-size:39px; line-height:.92; font-weight:900; letter-spacing:-0.045em; }} .label {{ margin-top:10px; color:var(--muted); font-size:12px; line-height:1.2; font-weight:740; }} .metric.dark .label {{ color:rgba(246,241,232,.74); }}
    .chart-panel {{ margin-top:24px; border:1.5px solid var(--line); border-radius:var(--radius); background:#11130f; color:var(--paper); padding:28px 24px 24px; height:2196px; display:grid; grid-template-rows:auto 1fr auto; gap:18px; }}
    .chart-head {{ display:flex; justify-content:space-between; gap:28px; align-items:end; border-bottom:1px solid rgba(246,241,232,.24); padding-bottom:16px; }} h2 {{ margin:0; font-size:38px; line-height:.95; letter-spacing:-0.015em; font-weight:630; }} .chart-sub {{ margin-top:8px; color:rgba(246,241,232,.7); font-size:14px; line-height:1.25; }}
    .legend {{ display:grid; grid-template-columns:auto auto auto auto; gap:8px 12px; align-items:center; color:rgba(246,241,232,.75); font-size:12px; font-weight:760; text-transform:uppercase; letter-spacing:.04em; white-space:nowrap; }} .swatch {{ width:15px; height:15px; border:1px solid rgba(246,241,232,.34); border-radius:999px; background:var(--lime); }} .swatch.orange {{ background:var(--orange); }} .swatch.blue {{ background:var(--blue); }} .swatch.pink {{ background:var(--pink); }} .swatch.gray {{ background:var(--gray); }}
    .treemap {{ position:relative; width:{TREEMAP_W}px; height:{TREEMAP_H}px; overflow:hidden; border-radius:18px; background:#191b16; border:1px solid rgba(246,241,232,.24); }}
    .tile {{ position:absolute; border:1px solid rgba(16,18,15,.42); border-radius:10px; overflow:hidden; padding:8px 9px; }} .tile.large {{ padding:10px 11px; border-radius:12px; }} .tile.medium {{ padding:8px 9px; border-radius:10px; }} .tile.small {{ padding:6px 7px; border-radius:8px; }} .tile.count {{ padding:0; border-radius:6px; }}
    .tile-copy {{ width:100%; max-width:100%; }} .tile.tier-top .tile-copy, .tile.tier-strong .tile-copy {{ width:50%; max-width:50%; }}
    .tile-rank {{ font-size:11px; line-height:1; font-weight:900; opacity:.65; letter-spacing:.08em; }} .tile-name {{ margin-top:5px; font-size:18px; line-height:.96; letter-spacing:-0.01em; font-weight:890; overflow-wrap:anywhere; }} .tile.medium .tile-name {{ font-size:14px; line-height:.98; }} .tile.small .tile-name {{ font-size:10.5px; line-height:1; }}
    .tile.tier-top .tile-rank {{ font-size:18px; }} .tile.tier-top .tile-name {{ margin-top:10px; font-size:42px; line-height:.92; letter-spacing:-0.025em; }} .tile.tier-top .tile-count {{ margin-top:12px; font-size:17px; }}
    .tile.tier-strong .tile-rank {{ font-size:15px; }} .tile.tier-strong .tile-name {{ margin-top:8px; font-size:31px; line-height:.93; letter-spacing:-0.02em; }} .tile.tier-strong .tile-count {{ margin-top:10px; font-size:14px; }}
    .tile-count {{ margin-top:6px; font-size:11px; line-height:1; font-weight:850; opacity:.74; }} .tile.medium .tile-count {{ font-size:9.5px; }} .tile.small .tile-count {{ font-size:8px; }} .count-only {{ position:absolute; inset:0; display:grid; place-items:center; font-size:9px; line-height:1; font-weight:900; opacity:.8; }}
    .note {{ color:rgba(246,241,232,.68); font-size:12px; line-height:1.25; font-weight:650; border-top:1px solid rgba(246,241,232,.22); padding-top:12px; }} .note strong {{ color:var(--paper); font-weight:900; }} .page-num {{ position:absolute; right:40px; bottom:24px; color:rgba(16,18,15,.38); font-size:11px; font-weight:800; letter-spacing:.16em; text-transform:uppercase; }}
    @page {{ size:12.5in 27.083333in; margin:0; }} @media print {{ body,*,*::before,*::after {{ -webkit-print-color-adjust:exact; print-color-adjust:exact; }} html,body {{ width:{SHEET_W}px; height:{SHEET_H}px; background:var(--paper); }} .sheet {{ width:{SHEET_W}px; height:{SHEET_H}px; background:var(--paper); }} }}
  </style>
</head>
<body>
  <main class="sheet">
    <section class="top">
      <div>
        <div class="eyebrow">ISPOR 2026 manufacturer data</div>
        <h1>Manufacturer title treemap</h1>
      </div>
      <div class="metrics" aria-label="summary metrics">
        <div class="metric dark"><div class="num">{nonblank:,}</div><div class="label">manufacturer attendees with populated normalized titles</div></div>
        <div class="metric"><div class="num">{len(items)}</div><div class="label">normalized title categories represented</div></div>
        <div class="metric"><div class="num">{pct(top20_total, nonblank)}</div><div class="label">of populated manufacturer titles in the top 20 categories</div></div>
        <div class="metric"><div class="num">{pct(director_type, nonblank)}</div><div class="label">Director-type titles: Director, Senior Director, Associate Director, and Executive Director</div></div>
      </div>
    </section>
    <section class="chart-panel">
      <div class="chart-head">
        <div>
          <h2>All normalized manufacturer titles</h2>
          <p class="chart-sub">Area represents attendee count; compact tiles show count labels when names would not fit.</p>
        </div>
        <div class="legend" aria-label="tile color legend">
          <span class="swatch"></span><span>Rank 1-3</span>
          <span class="swatch orange"></span><span>4-9</span>
          <span class="swatch blue"></span><span>10-20</span>
          <span class="swatch pink"></span><span>21-35</span>
        </div>
      </div>
      <div class="treemap" role="img" aria-label="Treemap of all normalized manufacturer titles by attendee count">
{tiles}
      </div>
      <p class="note"><strong>Manufacturer focus:</strong> Title normalization is available for {nonblank:,} of {total:,} manufacturer attendee rows. Percentages are calculated among those {nonblank:,} rows; all {len(items)} normalized title categories are represented, including {small_titles} categories with four or fewer rows.</p>
    </section>
    <div class="page-num">01 / 01</div>
  </main>
</body>
</html>
"""


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    total, blank, counts = read_counts()
    items = counts.most_common()
    write_csv(items, sum(counts.values()))
    HTML_OUT.write_text(build_html(total, blank, counts), encoding="utf-8")
    print(f"Wrote {HTML_OUT}")
    print(f"Wrote {CSV_OUT}")


if __name__ == "__main__":
    main()
