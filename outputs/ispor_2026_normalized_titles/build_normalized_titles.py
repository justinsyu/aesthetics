from collections import Counter
from html import escape
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Users/justinyu/Desktop/linkedin-posts")
INPUT = Path("/Users/justinyu/Downloads/ISPOR_2026_attendees (2).xlsx")
OUT_DIR = ROOT / "outputs" / "ispor_2026_normalized_titles"
HTML_OUT = OUT_DIR / "ispor_2026_normalized_titles.html"
CSV_OUT = OUT_DIR / "normalized_title_counts.csv"


def pct(value, total):
    return f"{round(value / total * 100)}%" if total else "0%"


def read_counts():
    wb = load_workbook(INPUT, read_only=True, data_only=True)
    ws = wb["Attendees"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    title_idx = headers.index("Normalized Title")
    affiliation_type_idx = headers.index("Affiliation Type")

    counts = Counter()
    blank = 0
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[affiliation_type_idx] or "").strip() != "Manufacturer":
            continue
        total += 1
        value = row[title_idx]
        title = "" if value is None else str(value).strip()
        if not title or title.lower() in {"nan", "none"}:
            blank += 1
            continue
        counts[title] += 1

    return total, blank, counts


def write_csv(counts, nonblank):
    lines = ["rank,normalized_title,count,share_of_populated_manufacturer_titles"]
    for rank, (title, count) in enumerate(counts.most_common(), 1):
        lines.append(
            f'{rank},"{title.replace(chr(34), chr(34) * 2)}",{count},{count / nonblank:.6f}'
        )
    CSV_OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_rows(top20, nonblank):
    max_count = top20[0][1]
    rows = []
    for rank, (title, count) in enumerate(top20, 1):
        width = round(count / max_count * 100, 1)
        share = count / nonblank * 100
        rows.append(
            '          <div class="bar-row">'
            f'<div class="rank">{rank:02d}</div>'
            f'<div class="title-label">{escape(title)}</div>'
            f'<div class="track"><div class="fill" style="width:{width}%"></div></div>'
            f'<div class="value">{count}</div>'
            f'<div class="pct">{share:.1f}%</div>'
            "</div>"
        )
    return "\n".join(rows)


def build_html(total, blank, counts):
    nonblank = sum(counts.values())
    top20 = counts.most_common(20)
    top20_total = sum(count for _, count in top20)
    unique = len(counts)
    director_family = sum(
        counts[title]
        for title in [
            "Director",
            "Senior Director",
            "Associate Director",
            "Executive Director",
        ]
    )
    manager_family = counts["Manager"] + counts["Senior Manager"]

    rows_html = build_rows(top20, nonblank)
    top_title, top_count = top20[0]

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>ISPOR 2026 Manufacturer Attendees: Normalized Title Mix</title>
  <style>
    :root {{
      --ink: #10120f;
      --muted: #5c6257;
      --paper: #f6f1e8;
      --paper-2: #ebe4d6;
      --card: #fffaf0;
      --line: #1b1f17;
      --lime: #d7ff5f;
      --orange: #ffb86b;
      --blue: #b8d8ff;
      --pink: #ffd3e0;
      --gray: #d6d0c2;
      --red: #ff8a76;
      --shadow: none;
      --radius: 26px;
    }}

    * {{ box-sizing: border-box; }}
    html, body {{ margin: 0; scrollbar-width: none; }}
    html::-webkit-scrollbar, body::-webkit-scrollbar {{ display: none; }}
    html {{ background: var(--paper); }}
    body {{
      background: var(--paper);
      color: var(--ink);
      font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      line-height: 1.28;
    }}

    .wrap {{
      width: min(1360px, calc(100vw - 48px));
      margin: 0 auto;
      height: 100%;
    }}

    .slide {{
      min-height: 100vh;
      height: 100vh;
      padding: 54px 0;
      position: relative;
      overflow: hidden;
      page-break-after: always;
      break-after: page;
      background: var(--paper);
    }}

    .layout {{
      display: grid;
      grid-template-columns: 430px 1fr;
      gap: 24px;
      height: 100%;
      align-items: stretch;
    }}

    .left {{
      display: grid;
      grid-template-rows: auto 1fr auto;
      gap: 18px;
      min-height: 0;
    }}

    .eyebrow {{
      display: inline-flex;
      align-items: center;
      width: fit-content;
      border: 1.4px solid var(--line);
      padding: 8px 11px;
      border-radius: 999px;
      font-size: 15px;
      font-weight: 850;
      letter-spacing: .06em;
      text-transform: uppercase;
      background: var(--lime);
      margin-bottom: 18px;
    }}

    h1 {{
      margin: 0;
      max-width: 420px;
      font-size: 70px;
      line-height: .9;
      letter-spacing: -0.04em;
      font-weight: 520;
    }}

    .dek {{
      margin-top: 17px;
      color: var(--muted);
      font-size: 20px;
      line-height: 1.25;
      max-width: 390px;
    }}

    .metrics {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 12px;
      align-self: end;
    }}

    .metric {{
      min-height: 116px;
      border: 1.5px solid var(--line);
      border-radius: var(--radius);
      background: rgba(255,250,240,.84);
      box-shadow: var(--shadow);
      padding: 16px;
    }}

    .metric.dark {{
      background: #11130f;
      color: var(--paper);
    }}

    .num {{
      font-size: 42px;
      line-height: .92;
      font-weight: 900;
      letter-spacing: -0.055em;
    }}

    .metric.dark .label {{ color: rgba(246,241,232,.74); }}

    .label {{
      margin-top: 11px;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.22;
      font-weight: 720;
    }}

    .note {{
      color: var(--muted);
      font-size: 14px;
      line-height: 1.25;
      max-width: 405px;
      border-top: 1px solid rgba(16,18,15,.2);
      padding-top: 11px;
    }}

    .note strong {{
      color: var(--ink);
      font-weight: 900;
    }}

    .chart-panel {{
      border: 1.5px solid var(--line);
      border-radius: var(--radius);
      background: #11130f;
      color: var(--paper);
      box-shadow: var(--shadow);
      padding: 26px 28px 24px;
      min-height: 0;
      display: grid;
      grid-template-rows: auto 1fr;
      gap: 18px;
    }}

    .chart-head {{
      display: flex;
      justify-content: space-between;
      gap: 26px;
      align-items: end;
      border-bottom: 1px solid rgba(246,241,232,.24);
      padding-bottom: 16px;
    }}

    h2 {{
      margin: 0;
      font-size: 36px;
      line-height: .95;
      letter-spacing: -0.035em;
      font-weight: 620;
    }}

    .chart-sub {{
      margin-top: 8px;
      color: rgba(246,241,232,.7);
      font-size: 14px;
      line-height: 1.25;
    }}

    .legend {{
      display: grid;
      grid-template-columns: auto auto;
      gap: 8px 12px;
      align-items: center;
      color: rgba(246,241,232,.75);
      font-size: 12px;
      font-weight: 760;
      text-transform: uppercase;
      letter-spacing: .04em;
      white-space: nowrap;
    }}

    .swatch {{
      width: 15px;
      height: 15px;
      border: 1px solid rgba(246,241,232,.34);
      border-radius: 999px;
      background: var(--lime);
    }}

    .swatch.orange {{ background: var(--orange); }}
    .swatch.blue {{ background: var(--blue); }}

    .bars {{
      display: grid;
      grid-template-rows: repeat(20, 1fr);
      gap: 7px;
      min-height: 0;
    }}

    .bar-row {{
      display: grid;
      grid-template-columns: 30px 220px 1fr 46px 54px;
      gap: 13px;
      align-items: center;
      min-height: 0;
    }}

    .rank {{
      color: rgba(246,241,232,.58);
      font-size: 13px;
      font-weight: 850;
      letter-spacing: .02em;
      text-align: right;
    }}

    .title-label {{
      color: var(--paper);
      font-size: 15px;
      line-height: 1.08;
      font-weight: 850;
      letter-spacing: -0.01em;
      overflow-wrap: anywhere;
    }}

    .track {{
      position: relative;
      height: 25px;
      border: 1px solid rgba(246,241,232,.24);
      border-radius: 999px;
      background: rgba(246,241,232,.09);
      overflow: hidden;
    }}

    .fill {{
      position: absolute;
      inset: 0 auto 0 0;
      min-width: 7px;
      border-radius: 999px;
      background: var(--blue);
    }}

    .bar-row:nth-child(-n+3) .fill {{ background: var(--lime); }}
    .bar-row:nth-child(4) .fill,
    .bar-row:nth-child(5) .fill,
    .bar-row:nth-child(6) .fill,
    .bar-row:nth-child(7) .fill,
    .bar-row:nth-child(8) .fill,
    .bar-row:nth-child(9) .fill {{ background: var(--orange); }}

    .value {{
      color: var(--paper);
      font-size: 15px;
      font-weight: 900;
      text-align: right;
    }}

    .pct {{
      color: rgba(246,241,232,.64);
      font-size: 13px;
      font-weight: 850;
      text-align: right;
    }}

    .slide-num {{
      position: absolute;
      right: 40px;
      bottom: 24px;
      color: rgba(16,18,15,.38);
      font-size: 11px;
      font-weight: 800;
      letter-spacing: .16em;
      text-transform: uppercase;
    }}

    @page {{ size: 16.6666667in 9.375in; margin: 0; }}
    @media print {{
      body, *, *::before, *::after {{
        -webkit-print-color-adjust: exact;
        print-color-adjust: exact;
      }}
      html, body {{
        width: 1600px;
        height: 900px;
        background: var(--paper);
      }}
      .slide {{
        width: 1600px;
        height: 900px;
        min-height: 900px;
        padding: 54px 0;
        background: var(--paper);
      }}
      .wrap {{ width: 1360px; height: 100%; }}
    }}
  </style>
</head>
<body>
  <article class="slide">
    <div class="wrap layout">
      <section class="left">
        <div>
          <div class="eyebrow">ISPOR 2026 manufacturer data</div>
          <h1>Manufacturer attendee title profile</h1>
          <p class="dek">Rows ranked by normalized attendee title among manufacturer attendees.</p>
        </div>

        <div class="metrics" aria-label="summary metrics">
          <div class="metric dark">
            <div class="num">{nonblank:,}</div>
            <div class="label">manufacturer attendees with populated normalized titles</div>
          </div>
          <div class="metric">
            <div class="num">{unique}</div>
            <div class="label">unique normalized title categories</div>
          </div>
          <div class="metric">
            <div class="num">{pct(top20_total, nonblank)}</div>
            <div class="label">of populated manufacturer titles in the top 20 categories</div>
          </div>
          <div class="metric">
            <div class="num">{pct(director_family, nonblank)}</div>
            <div class="label">Director-type titles: Director, Senior Director, Associate Director, and Executive Director</div>
          </div>
        </div>

        <p class="note"><strong>Manufacturer focus:</strong> Title normalization is available for {nonblank:,} of {total:,} manufacturer attendee rows. Percentages are calculated among those {nonblank:,} rows; the largest single title is {escape(top_title)} at {top_count:,} rows.</p>
      </section>

      <section class="chart-panel">
        <div class="chart-head">
          <div>
            <h2>Manufacturer attendees by title, top 20</h2>
            <p class="chart-sub">Sorted descending. Right columns show row count and share of populated manufacturer titles.</p>
          </div>
          <div class="legend" aria-label="bar color legend">
            <span class="swatch"></span><span>Top 3</span>
            <span class="swatch orange"></span><span>Rank 4-9</span>
            <span class="swatch blue"></span><span>Rank 10-20</span>
          </div>
        </div>

        <div class="bars" role="img" aria-label="Horizontal bar chart of the top 20 normalized titles by attendee count">
{rows_html}
        </div>
      </section>
    </div>
    <div class="slide-num">01 / 01</div>
  </article>
</body>
</html>
"""


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    total, blank, counts = read_counts()
    nonblank = sum(counts.values())
    write_csv(counts, nonblank)
    HTML_OUT.write_text(build_html(total, blank, counts), encoding="utf-8")
    print(f"Wrote {HTML_OUT}")
    print(f"Wrote {CSV_OUT}")


if __name__ == "__main__":
    main()
